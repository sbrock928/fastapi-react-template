"""API router for the reporting module - Phase 1: Fixed async/sync issues."""

from typing import List, Dict, Any, Optional
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
import pandas as pd
import io
from app.reporting.service import ReportService
from app.reporting.dao import ReportDAO
from app.reporting.schemas import (
    ReportRead,
    ReportCreate,
    ReportUpdate,
    ReportSummary,
    RunReportRequest,
    AvailableCalculation,
    ReportScope,
)
from app.core.dependencies import SessionDep, DWSessionDep, get_user_calculation_service, get_system_calculation_service, get_report_execution_service, get_cdi_calculation_service
from app.datawarehouse.dao import DatawarehouseDAO
from app.calculations.service import UserCalculationService, SystemCalculationService, ReportExecutionService
from app.calculations.cdi_service import CDIVariableCalculationService
import io
import pandas as pd
from fastapi import Response, HTTPException
from typing import Dict, Any

router = APIRouter(prefix="/reports", tags=["reporting"])


# Dependency functions
def get_report_dao(db: SessionDep) -> ReportDAO:
    return ReportDAO(db)


def get_dw_dao(db: DWSessionDep) -> DatawarehouseDAO:
    return DatawarehouseDAO(db)


def get_report_service(
    report_dao: ReportDAO = Depends(get_report_dao),
    dw_dao: DatawarehouseDAO = Depends(get_dw_dao),
    user_calc_service: UserCalculationService = Depends(get_user_calculation_service),
    system_calc_service: SystemCalculationService = Depends(get_system_calculation_service),
    report_execution_service: ReportExecutionService = Depends(get_report_execution_service)
) -> ReportService:
    return ReportService(
        report_dao, 
        dw_dao, 
        user_calc_service, 
        system_calc_service, 
        report_execution_service
    )


# ===== REPORT CONFIGURATION ENDPOINTS =====


@router.get("/", response_model=List[ReportRead])
async def get_all_reports(service: ReportService = Depends(get_report_service)) -> List[ReportRead]:
    """Get all report configurations."""
    return await service.get_all()  # FIXED: added await


@router.get("/summary", response_model=List[ReportSummary])
async def get_all_reports_summary(
    service: ReportService = Depends(get_report_service),
) -> List[ReportSummary]:
    """Get all reports with summary information."""
    return await service.get_all_summaries()  # FIXED: added await


@router.get("/{report_id}", response_model=ReportRead)
async def get_report_by_id(
    report_id: int, service: ReportService = Depends(get_report_service)
) -> ReportRead:
    """Get a specific report configuration by ID."""
    report = await service.get_by_id(report_id)  # FIXED: added await
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    return report


@router.post("/", response_model=ReportRead)
async def create_report(
    report_data: ReportCreate, service: ReportService = Depends(get_report_service)
) -> ReportRead:
    """Create a new report configuration."""
    return await service.create(report_data)  # FIXED: added await


@router.patch("/{report_id}", response_model=ReportRead)
async def update_report(
    report_id: int, report_data: ReportUpdate, service: ReportService = Depends(get_report_service)
) -> ReportRead:
    """Update an existing report configuration."""
    report = await service.update(report_id, report_data)  # FIXED: added await
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    return report


@router.delete("/{report_id}")
async def delete_report(
    report_id: int, service: ReportService = Depends(get_report_service)
) -> Dict[str, str]:
    """Delete a report configuration."""
    success = await service.delete(report_id)  # FIXED: added await
    if not success:
        raise HTTPException(status_code=404, detail="Report not found")
    return {"message": "Report deleted successfully"}


# ===== REPORT EXECUTION ENDPOINTS =====


@router.post("/run", response_model=List[Dict[str, Any]])
async def run_report(
    request: RunReportRequest, service: ReportService = Depends(get_report_service)
) -> List[Dict[str, Any]]:
    """Run a saved report configuration."""
    return await service.run_report(
        request.report_id, request.cycle_code
    )  # FIXED: changed from run_saved_report to run_report


@router.post("/run/{report_id}")
async def run_report_by_id(
    report_id: int, request: Dict[str, Any], service: ReportService = Depends(get_report_service)
) -> Dict[str, Any]:
    """Run a saved report by ID with cycle parameter - returns data with column metadata."""
    cycle_code = request.get("cycle_code")
    if not cycle_code:
        raise HTTPException(status_code=400, detail="cycle_code is required")
    
    # Get the report configuration to extract column metadata
    report = await service.get_by_id(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Run the report to get formatted data
    data = await service.run_report(report_id, cycle_code)
    
    # Extract column metadata from the report's column preferences
    column_metadata = []
    if report.column_preferences and report.column_preferences.columns:
        for col_pref in report.column_preferences.columns:
            if col_pref.is_visible:
                column_metadata.append({
                    "field": col_pref.display_name,  # Use display name as field (matches data keys)
                    "header": col_pref.display_name,
                    "format_type": col_pref.format_type,
                    "display_order": col_pref.display_order
                })
    else:
        # Fallback: create basic metadata from data keys if no column preferences
        if data:
            for i, key in enumerate(data[0].keys()):
                column_metadata.append({
                    "field": key,
                    "header": key,
                    "format_type": "text",  # Default to text formatting
                    "display_order": i
                })
    
    # Sort columns by display order
    column_metadata.sort(key=lambda x: x.get("display_order", 0))
    
    return {
        "data": data,
        "columns": column_metadata,
        "total_rows": len(data)
    }


# ===== PREVIEW AND EXECUTION LOG ENDPOINTS =====


@router.get("/{report_id}/structure")
async def get_report_structure(
    report_id: int, service: ReportService = Depends(get_report_service)
) -> Dict[str, Any]:
    """Get report structure information for skeleton mode display."""
    report = await service.get_by_id(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Extract column structure from column preferences
    columns = []
    if report.column_preferences and report.column_preferences.columns:
        visible_columns = [col for col in report.column_preferences.columns if col.is_visible]
        visible_columns.sort(key=lambda x: x.display_order)
        
        for col_pref in visible_columns:
            columns.append({
                "field": col_pref.display_name.lower().replace(' ', '_'),
                "header": col_pref.display_name,
                "format_type": col_pref.format_type,
                "display_order": col_pref.display_order
            })
    else:
        # Fallback structure if no column preferences
        default_columns = [
            {"field": "deal_number", "header": "Deal Number", "format_type": "number", "display_order": 0},
            {"field": "cycle_code", "header": "Cycle Code", "format_type": "number", "display_order": 1}
        ]
        
        if report.scope == "TRANCHE":
            default_columns.insert(1, {
                "field": "tranche_id", "header": "Tranche ID", "format_type": "text", "display_order": 1
            })
            # Adjust other display orders
            for i, col in enumerate(default_columns[2:], 2):
                col["display_order"] = i
        
        columns = default_columns
    
    return {
        "report_id": report_id,
        "name": report.name,
        "description": report.description,
        "scope": report.scope,
        "columns": columns,
        "deal_count": len(report.selected_deals) if report.selected_deals else 0,
        "calculation_count": len(report.selected_calculations) if report.selected_calculations else 0
    }


@router.get("/{report_id}/preview-sql")
async def preview_report_sql(
    report_id: int, cycle_code: int = 202404, service: ReportService = Depends(get_report_service)
) -> Dict[str, Any]:
    """Preview SQL that would be generated for a report."""
    return await service.preview_report_sql(report_id, cycle_code)  # FIXED: added await


@router.get("/{report_id}/execution-logs")
async def get_report_execution_logs(
    report_id: int, limit: int = 50, service: ReportService = Depends(get_report_service)
) -> List[Dict[str, Any]]:
    """Get execution logs for a report."""
    return await service.get_execution_logs(report_id, limit)  # FIXED: added await


# ===== CALCULATION CONFIGURATION ENDPOINTS =====


@router.get("/calculations/available/{scope}", response_model=List[AvailableCalculation])
def get_available_calculations(
    scope: str, service: ReportService = Depends(get_report_service)
) -> List[AvailableCalculation]:
    """Get available calculations for report configuration based on scope with new calculation_id format"""
    try:
        scope_enum = ReportScope(scope.upper())
        calculations = service.get_available_calculations_for_scope(scope_enum)
        
        # Add debugging info about the new format
        print(f"Debug: Returning {len(calculations)} calculations with new format:")
        for calc in calculations[:3]:  # Print first 3 for debugging
            print(f"  - ID: {calc.id}, Name: {calc.name}, Type: {calc.calculation_type}")
        
        return calculations
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid scope: {scope}")


# ===== DATA ENDPOINTS (for report building) =====


@router.get("/data/issuer-codes", response_model=List[str])
def get_available_issuer_codes(
    service: ReportService = Depends(get_report_service),
) -> List[str]:
    """Get unique issuer codes for deal filtering."""
    deals = service.get_available_deals()  # This now returns dictionaries
    # Fix: Use dictionary access since service returns dictionaries
    issuer_codes = sorted(list(set(deal["issr_cde"] for deal in deals)))
    return issuer_codes


@router.get("/data/deals", response_model=List[Dict[str, Any]])
def get_available_deals(
    issuer_code: Optional[str] = None,
    service: ReportService = Depends(get_report_service),
) -> List[Dict[str, Any]]:
    """Get available deals for report building, optionally filtered by issuer code."""
    deals = service.get_available_deals()  # This now returns dictionaries

    # Filter by issuer code if provided
    if issuer_code:
        # Fix: Use dictionary access since service returns dictionaries
        deals = [deal for deal in deals if deal["issr_cde"] == issuer_code]

    return deals


@router.post("/data/tranches", response_model=Dict[int, List[Dict[str, Any]]])
def get_available_tranches(
    request: Dict[str, Any], service: ReportService = Depends(get_report_service)
) -> Dict[int, List[Dict[str, Any]]]:
    """Get available tranches for specific deals."""
    deal_ids = request.get("dl_nbrs", [])
    cycle_code = request.get("cycle_code")

    tranches_by_deal = service.get_available_tranches_for_deals(
        deal_ids, cycle_code
    )  # This one stays sync
    return {int(deal_id): tranches for deal_id, tranches in tranches_by_deal.items()}


@router.get("/data/deals/{dl_nbr}/tranches", response_model=List[Dict[str, Any]])
def get_deal_tranches(
    dl_nbr: int, 
    cycle_code: Optional[int] = None,
    service: ReportService = Depends(get_report_service)
) -> List[Dict[str, Any]]:
    """Get available tranches for a specific deal."""
    tranches_by_deal = service.get_available_tranches_for_deals(
        [dl_nbr], cycle_code
    )
    return tranches_by_deal.get(dl_nbr, [])


@router.get("/data/cycles", response_model=List[Dict[str, Any]])
def get_available_cycles(
    service: ReportService = Depends(get_report_service),
) -> List[Dict[str, str]]:
    """Get available cycle codes from the data warehouse."""
    return service.get_available_cycles()  # This one stays sync


# ===== EXPORT ENDPOINTS =====


@router.post("/export-xlsx")
async def export_to_xlsx(request: Dict[str, Any]) -> Response:
    """
    Export report data to Excel (XLSX) format with column management support.
    The data should already be formatted and filtered by the backend.
    """
    try:
        report_type = request.get("reportType", "Report")
        data = request.get("data", [])
        file_name = request.get("fileName", "report.xlsx")
        
        # Optional: Include column preferences for additional formatting
        column_preferences = request.get("columnPreferences")

        if not data:
            raise HTTPException(status_code=400, detail="No data provided for export")

        # Create DataFrame from the pre-formatted data
        df = pd.DataFrame(data)
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Data is empty")

        # Create Excel file in memory
        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # Write the data to the Excel file
            sheet_name = report_type[:31]  # Excel sheet name limit is 31 chars
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Apply additional formatting based on column preferences
            if column_preferences:
                _apply_excel_formatting(worksheet, df, column_preferences)
            else:
                # Default formatting - auto-adjust column widths
                _auto_adjust_columns(worksheet)

        excel_buffer.seek(0)

        # Ensure the filename has .xlsx extension
        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"

        # Return the Excel file as a response
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={file_name}"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating Excel file: {str(e)}")


def _apply_excel_formatting(worksheet, df: pd.DataFrame, column_preferences: Dict[str, Any]):
    """Apply Excel formatting based on column preferences."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Column-specific formatting based on preferences
    columns_info = column_preferences.get("columns", [])
    
    for col_info in columns_info:
        if not col_info.get("is_visible", True):
            continue
            
        display_name = col_info.get("display_name", "")
        format_type = col_info.get("format_type", "text")
        
        # Find column index by display name
        try:
            col_index = df.columns.get_loc(display_name) + 1
            col_letter = get_column_letter(col_index)
            
            # Apply format-specific styling
            if format_type == "currency":
                # Format currency columns
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_index)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('$'):
                        cell.alignment = Alignment(horizontal="right")
                        
            elif format_type == "percentage":
                # Format percentage columns  
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_index)
                    if cell.value and isinstance(cell.value, str) and cell.value.endswith('%'):
                        cell.alignment = Alignment(horizontal="center")
                        
            elif format_type == "number":
                # Format number columns
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_index)
                    cell.alignment = Alignment(horizontal="right")
                    
            elif format_type in ["date_mdy", "date_dmy"]:
                # Format date columns
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_index)
                    cell.alignment = Alignment(horizontal="center")
        
        except (KeyError, ValueError):
            # Column not found, skip formatting
            continue

    # Auto-adjust column widths
    _auto_adjust_columns(worksheet)


def _auto_adjust_columns(worksheet):
    """Auto-adjust column widths for better readability."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        # Set column width with some padding, max 50 characters
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width